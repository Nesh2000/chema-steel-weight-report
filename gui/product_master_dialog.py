"""
Product Master Module for Steel Weight Calculator.
Admin interface for managing the product database, including weights and aliases.
Supports import/export from/to Excel (using openpyxl).
"""

import logging
from pathlib import Path
from typing import Optional, Dict

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QGridLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QLineEdit, QDoubleSpinBox, QMessageBox, QFileDialog, QHeaderView
)
from PySide6.QtCore import Qt

logger = logging.getLogger(__name__)


try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. Excel import/export features will be disabled.")


class ProductMasterDialog(QDialog):
    """Dialog for managing steel products in the master database."""

    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Product Master Data")
        self.setMinimumSize(1000, 700)
        
        self.db_manager = db_manager
        self.products = []
        
        self.init_ui()
        self.refresh_data()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Top Controls
        top_layout = QHBoxLayout()
        self.input_search = QLineEdit()
        self.input_search.setPlaceholderText("Search products...")
        self.btn_search = QPushButton("Search")
        self.btn_search.clicked.connect(self.search_products)
        
        self.btn_add = QPushButton("Add Product")
        self.btn_add.clicked.connect(self.add_product)
        
        self.btn_import = QPushButton("Import from Excel")
        self.btn_import.clicked.connect(self.import_excel)
        
        self.btn_import_inventory = QPushButton("Import Inventory Excel")
        self.btn_import_inventory.clicked.connect(self.import_inventory_excel)
        
        self.btn_export = QPushButton("Export to Excel")
        self.btn_export.clicked.connect(self.export_excel)
        
        top_layout.addWidget(QLabel("Search:"))
        top_layout.addWidget(self.input_search)
        top_layout.addWidget(self.btn_search)
        top_layout.addWidget(self.btn_add)
        top_layout.addWidget(self.btn_import)
        top_layout.addWidget(self.btn_import_inventory)
        top_layout.addWidget(self.btn_export)
        layout.addLayout(top_layout)
        
        # Products Table
        self.products_table = QTableWidget()
        self.products_table.setColumnCount(11)
        self.products_table.setHorizontalHeaderLabels([
            "ID", "Product Name", "Category", "Size", "Diameter (mm)", 
            "Thickness (mm)", "Width (mm)", "Length (m)", "Unit Weight (kg)", 
            "Aliases", "Status"
        ])
        self.products_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.products_table.setSelectionBehavior(QTableWidget.SelectRows)
        layout.addWidget(self.products_table)
        
        # Bottom Buttons
        bottom_layout = QHBoxLayout()
        self.btn_edit = QPushButton("Edit Selected")
        self.btn_edit.clicked.connect(self.edit_selected_product)
        
        self.btn_delete = QPushButton("Delete Selected")
        self.btn_delete.clicked.connect(self.delete_selected_product)
        
        self.btn_close = QPushButton("Close")
        self.btn_close.clicked.connect(self.reject)
        
        bottom_layout.addWidget(self.btn_edit)
        bottom_layout.addWidget(self.btn_delete)
        bottom_layout.addStretch()
        bottom_layout.addWidget(self.btn_close)
        layout.addLayout(bottom_layout)

    def refresh_data(self):
        self.products = self.db_manager.get_all_products()
        self.populate_table()
        
    def populate_table(self):
        self.products_table.setRowCount(len(self.products))
        for row, product in enumerate(self.products):
            self.products_table.setItem(row, 0, QTableWidgetItem(str(product["product_id"])))
            self.products_table.setItem(row, 1, QTableWidgetItem(product["product_name"]))
            self.products_table.setItem(row, 2, QTableWidgetItem(product.get("category", "")))
            self.products_table.setItem(row, 3, QTableWidgetItem(product.get("size", "") or ""))
            
            # Handle None values for numeric fields
            self.products_table.setItem(row, 4, QTableWidgetItem(str(product.get("diameter_mm") or "-")))
            self.products_table.setItem(row, 5, QTableWidgetItem(str(product.get("thickness_mm") or "-")))
            self.products_table.setItem(row, 6, QTableWidgetItem(str(product.get("width_mm") or "-")))
            self.products_table.setItem(row, 7, QTableWidgetItem(str(product.get("length_m") or "-")))
            self.products_table.setItem(row, 8, QTableWidgetItem(str(product["unit_weight_kg"])))
            self.products_table.setItem(row, 9, QTableWidgetItem(product.get("aliases", "")))
            self.products_table.setItem(row, 10, QTableWidgetItem("Active" if product.get("is_active", 1) else "Inactive"))

    def search_products(self):
        """Search for products and update the table."""
        search_term = self.input_search.text()
        if not search_term:
            self.refresh_data()
            return
        self.products = self.db_manager.search_products(search_term)
        self.populate_table()

    def add_product(self):
        """Open a dialog to add a new product."""
        dialog = ProductEditDialog(self.db_manager, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.refresh_data()

    def edit_selected_product(self):
        """Open dialog to edit the selected product."""
        selected_row = self.products_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Warning", "Please select a product to edit.")
            return
        
        product_id = int(self.products_table.item(selected_row, 0).text())
        dialog = ProductEditDialog(self.db_manager, product_id=product_id, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.refresh_data()

    def delete_selected_product(self):
        """Delete the selected product after confirmation."""
        selected_row = self.products_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Warning", "Please select a product to delete.")
            return
        
        product_id = int(self.products_table.item(selected_row, 0).text())
        product_name = self.products_table.item(selected_row, 1).text()
        
        reply = QMessageBox.question(self, "Confirm Delete", 
                                     f"Are you sure you want to delete the product '{product_name}'?\n\n"
                                     "This action cannot be undone.")
        
        if reply == QMessageBox.StandardButton.Yes:
            self.db_manager.delete_product(product_id)
            self.refresh_data()
            QMessageBox.information(self, "Deleted", f"Product '{product_name}' has been deleted.")

    def import_excel(self):
        """Import products from an Excel file."""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Error", "openpyxl library is not installed. Please install it to use this feature.")
            return
            
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)")
        if not file_name:
            return
            
        try:
            wb = load_workbook(file_name)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            # Mappings from header to DB keys
            key_map = {
                "product_name": ["product_name", "name", "product", "item_name"],
                "category": ["category", "cat", "type"],
                "size": ["size", "dimension"],
                "diameter_mm": ["diameter_mm", "diameter", "dia_mm", "dia"],
                "thickness_mm": ["thickness_mm", "thickness", "thk_mm", "thk"],
                "width_mm": ["width_mm", "width", "w_mm"],
                "length_m": ["length_m", "length", "len", "len_m"],
                "unit_weight_kg": ["unit_weight_kg", "unit_weight", "weight_kg", "weight", "wt_kg"],
                "aliases": ["aliases", "alias", "alternative_names", "alt_names"]
            }
            
            column_map = {}
            for db_key, aliases in key_map.items():
                for col_idx, header in enumerate(headers):
                    if header and any(alias.lower() in str(header).lower() for alias in aliases):
                        column_map[db_key] = col_idx
                        break
            
            imported_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                
                product_data = {}
                for db_key, col_idx in column_map.items():
                    if col_idx < len(row):
                        product_data[db_key] = row[col_idx]
                        
                if product_data:
                    self.db_manager.add_product(product_data)
                    imported_count += 1
            
            self.refresh_data()
            QMessageBox.information(self, "Import Success", f"Successfully imported {imported_count} products from the Excel file.")
            
        except Exception as e:
            logger.error(f"Error importing Excel file: {e}")
            QMessageBox.critical(self, "Import Error", f"Failed to import file:\n{e}")

    def import_inventory_excel(self):
        """Import products from a fixed-layout inventory Excel file."""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Error", "openpyxl library is not installed. Please install it to use this feature.")
            return

        file_name, _ = QFileDialog.getOpenFileName(self, "Open Inventory Excel File", "", "Excel Files (*.xlsx)")
        if not file_name:
            return

        reply = QMessageBox.question(
            self, "Confirm Import",
            "This will add all inventory items from the selected Excel file. "
            "Existing products will not be deleted. Continue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            wb = load_workbook(file_name, data_only=True)
            ws = None
            if "Inventory" in wb.sheetnames:
                ws = wb["Inventory"]
            else:
                ws = wb.active

            imported_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 7:
                    continue
                description = row[1]
                if not description or str(description).strip() == "":
                    continue
                description = str(description).strip()
                # Derive category from description
                category = "General"
                desc_upper = description.upper()
                if desc_upper.startswith("TUBE"):
                    category = "Tube"
                elif desc_upper.startswith("ANGLE"):
                    category = "Angle"
                elif desc_upper.startswith("FLAT BAR"):
                    category = "Flat Bar"
                elif desc_upper.startswith("MILD STEEL PLATE") or desc_upper.startswith("MS PLATE") or desc_upper.startswith("CHEQUERED"):
                    category = "Sheet/Plate"
                elif desc_upper.startswith("BLACK PIPE"):
                    category = "Black Pipe"
                elif desc_upper.startswith("ROUND FURNITURE") or desc_upper.startswith("ROUND FURN"):
                    category = "Round Furniture Pipe"
                elif "ZED" in desc_upper:
                    category = "Zed"
                elif "BRC" in desc_upper or "WIREMESH" in desc_upper:
                    category = "Mesh"
                try:
                    unit_weight_kg = float(row[6]) if row[6] is not None else 0.0
                except (ValueError, TypeError):
                    unit_weight_kg = 0.0
                count_val = str(row[4]).strip() if row[4] is not None else ""
                is_active = 0 if count_val == "Discontinued" else 1
                product_data = {
                    "product_name": description,
                    "category": category,
                    "unit_weight_kg": unit_weight_kg,
                    "aliases": "",
                    "is_active": is_active,
                }
                self.db_manager.add_product(product_data)
                imported_count += 1

            self.refresh_data()
            QMessageBox.information(self, "Import Success", f"Successfully imported {imported_count} products from the inventory file.")

        except Exception as e:
            logger.error(f"Error importing inventory Excel file: {e}")
            QMessageBox.critical(self, "Import Error", f"Failed to import file:\n{e}")

    def export_excel(self):
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Error", "openpyxl library is not installed. Please install it to use this feature.")
            return
            
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "products_export.xlsx", "Excel Files (*.xlsx)")
        if not file_name:
            return
            
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # Header
            headers = ["ID", "Product Name", "Category", "Size", "Diameter (mm)", 
                       "Thickness (mm)", "Width (mm)", "Length (m)", "Unit Weight (kg)", "Aliases"]
            ws.append(headers)
            
            # Data
            for product in self.db_manager.get_all_products():
                ws.append([
                    product["product_id"], product["product_name"], product.get("category"), 
                    product.get("size"), product.get("diameter_mm"), product.get("thickness_mm"),
                    product.get("width_mm"), product.get("length_m"), product["unit_weight_kg"], 
                    product.get("aliases", "")
                ])
            
            # Styling (optional)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(file_name)
            QMessageBox.information(self, "Export Success", "Products successfully exported to Excel.")
            
        except Exception as e:
            logger.error(f"Error exporting Excel file: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to export file:\n{e}")


class ProductEditDialog(QDialog):
    """Dialog for adding or editing a single product."""

    def __init__(self, db_manager, product_id: Optional[int] = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Product" if product_id else "Add Product")
        self.setMinimumSize(400, 400)
        
        self.db_manager = db_manager
        self.product_id = product_id
        self.is_edit_mode = product_id is not None
        
        self.init_ui()
        
        if self.is_edit_mode:
            self.load_product_data()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        grid_layout = QGridLayout()
        
        # Row 0: Product Name
        grid_layout.addWidget(QLabel("Product Name: *"), 0, 0)
        self.input_name = QLineEdit()
        grid_layout.addWidget(self.input_name, 0, 1)
        
        # Row 1: Category
        grid_layout.addWidget(QLabel("Category:"), 1, 0)
        self.input_category = QLineEdit()
        grid_layout.addWidget(self.input_category, 1, 1)
        
        # Row 2: Size
        grid_layout.addWidget(QLabel("Size:"), 2, 0)
        self.input_size = QLineEdit()
        grid_layout.addWidget(self.input_size, 2, 1)
        
        # Row 3: Diameter
        grid_layout.addWidget(QLabel("Diameter (mm):"), 3, 0)
        self.input_diameter = QDoubleSpinBox()
        self.input_diameter.setDecimals(2)
        self.input_diameter.setRange(0, 99999)
        grid_layout.addWidget(self.input_diameter, 3, 1)
        
        # Row 4: Thickness
        grid_layout.addWidget(QLabel("Thickness (mm):"), 4, 0)
        self.input_thickness = QDoubleSpinBox()
        self.input_thickness.setDecimals(2)
        self.input_thickness.setRange(0, 99999)
        grid_layout.addWidget(self.input_thickness, 4, 1)
        
        # Row 5: Width
        grid_layout.addWidget(QLabel("Width (mm):"), 5, 0)
        self.input_width = QDoubleSpinBox()
        self.input_width.setDecimals(2)
        self.input_width.setRange(0, 99999)
        grid_layout.addWidget(self.input_width, 5, 1)
        
        # Row 6: Length
        grid_layout.addWidget(QLabel("Length (m):"), 6, 0)
        self.input_length = QDoubleSpinBox()
        self.input_length.setDecimals(2)
        self.input_length.setRange(0, 99999)
        grid_layout.addWidget(self.input_length, 6, 1)
        
        # Row 7: Unit Weight
        grid_layout.addWidget(QLabel("Unit Weight (kg): *"), 7, 0)
        self.input_weight = QDoubleSpinBox()
        self.input_weight.setDecimals(3)
        self.input_weight.setRange(0, 9999999.999)
        grid_layout.addWidget(self.input_weight, 7, 1)
        
        # Row 8: Aliases
        grid_layout.addWidget(QLabel("Aliases (comma-separated):"), 8, 0)
        self.input_aliases = QLineEdit()
        self.input_aliases.setToolTip("e.g., Y12, D12, 12MM TMT")
        grid_layout.addWidget(self.input_aliases, 8, 1)
        
        layout.addLayout(grid_layout)
        
        # Buttons
        btn_layout = QHBoxLayout()
        self.btn_save = QPushButton("Save")
        self.btn_save.clicked.connect(self.save_product)
        
        self.btn_cancel = QPushButton("Cancel")
        self.btn_cancel.clicked.connect(self.reject)
        
        btn_layout.addWidget(self.btn_save)
        btn_layout.addWidget(self.btn_cancel)
        layout.addLayout(btn_layout)

    def load_product_data(self):
        product = self.db_manager.get_product_by_id(self.product_id)
        if not product:
            return
        
        self.input_name.setText(product["product_name"])
        self.input_category.setText(product.get("category", ""))
        self.input_size.setText(product.get("size", ""))
        
        self.input_diameter.setValue(product.get("diameter_mm") or 0)
        self.input_thickness.setValue(product.get("thickness_mm") or 0)
        self.input_width.setValue(product.get("width_mm") or 0)
        self.input_length.setValue(product.get("length_m") or 0)
        self.input_weight.setValue(product.get("unit_weight_kg", 0))
        self.input_aliases.setText(product.get("aliases", ""))

    def save_product(self):
        if not self.input_name.text().strip():
            QMessageBox.warning(self, "Validation Error", "Product Name is required.")
            return
        
        product_data = {
            "product_name": self.input_name.text().strip(),
            "category": self.input_category.text().strip(),
            "size": self.input_size.text().strip(),
            "diameter_mm": self.input_diameter.value() if self.input_diameter.value() > 0 else None,
            "thickness_mm": self.input_thickness.value() if self.input_thickness.value() > 0 else None,
            "width_mm": self.input_width.value() if self.input_width.value() > 0 else None,
            "length_m": self.input_length.value() if self.input_length.value() > 0 else None,
            "unit_weight_kg": self.input_weight.value(),
            "aliases": self.input_aliases.text().strip()
        }
        
        try:
            if self.is_edit_mode:
                self.db_manager.update_product(self.product_id, product_data)
            else:
                self.db_manager.add_product(product_data)
            
            QMessageBox.information(self, "Success", "Product saved successfully.")
            self.accept()
        except Exception as e:
            logger.error(f"Error saving product: {e}")
            QMessageBox.critical(self, "Error", f"Failed to save product:\n{e}")
